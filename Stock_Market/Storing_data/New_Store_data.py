from Access_Data.Data import Recived_Data
import openpyxl
# import pandas as pd

data = Recived_Data()
Symbol = input("Enter the Stock Symbol : ")
Series = input("Enter the Stock Series : ")
Year = int(input("Enter the Year : "))
Month = int(input("Enter the Month : "))
Day = int(input("Enter the Day : "))
Quantity = int(input("Enter the No_Of_Shares : "))


New_Data = data.Get_Historical_data_Particular_Day(Symbol, Series, Year, Month, Day)
excel_path = "historical_data_particular_day.xlsx"
New_Data.to_excel(excel_path, index=False)

# excel_file_path = "Result_data.xlsx"
# wb1 = openpyxl.load_workbook(excel_file_path)

wb = openpyxl.load_workbook(excel_path)
source_worksheet = wb['Sheet1']

data_to_add = []
for source_row in source_worksheet.iter_rows(min_row=2, values_only=True):
    date = source_row[0]
    series = source_row[1]
    value = source_row[7]

    data_to_add.append(tuple([date, series, value]))


# Adding extra column headers
extra_column_headers = ['Quantity', 'Invested_Amount', 'Profit', 'Profit (%)']
for col_index, header in enumerate(extra_column_headers, start=source_worksheet.max_column + 1):
    source_worksheet.cell(row=1, column=col_index, value=header)

# Write data and add values for extra columns
for row_index, data_row in enumerate(data_to_add, start=1):
    date, series, value = data_row

    # Write data to existing columns
    source_worksheet.cell(row=row_index, column=1, value=date)
    source_worksheet.cell(row=row_index, column=2, value=series)
    source_worksheet.cell(row=row_index, column=3, value=value)

    # Add values for extra columns
    source_worksheet.cell(row=row_index, column=4, value='Quantity')
    source_worksheet.cell(row=row_index, column=4, value='Invested_Amount')
    source_worksheet.cell(row=row_index, column=4, value='Profit')
    source_worksheet.cell(row=row_index, column=4, value='Profit (%)')

wb.save(excel_path)
wb.close()
print("Updated Successfully")


# excel_file_path = "Result_data.xlsx"
# wb = openpyxl.load_workbook(excel_file_path)
